"""Case service — load and manage patient cases from Excel."""

import re
from pathlib import Path

import openpyxl

# ── Department keyword matching (from app.py) ──

DEPARTMENT_KEYWORDS = {
    "急诊科": ["外伤", "车祸", "坠落", "摔伤", "出血", "晕厥", "休克", "中毒", "烧伤", "溺水", "电击", "窒息", "过敏", "呼吸困难", "心跳", "猝死", "急性", "抢救", "急救", "昏迷"],
    "内科": ["高血压", "糖尿病", "冠心病", "心衰", "肺炎", "哮喘", "肝炎", "肝硬化", "肾炎", "贫血", "白血病", "甲亢", "痛风", "关节炎", "胃痛", "腹泻", "便秘", "头晕", "乏力", "消瘦", "发热", "咳嗽", "胸闷", "心悸", "水肿", "黄疸", "恶心", "呕吐", "腹胀"],
    "外科": ["骨折", "脱位", "扭伤", "拉伤", "刀伤", "烧伤疤痕", "脓肿", "痔疮", "疝气", "阑尾炎", "胆囊炎", "胰腺炎", "肠梗阻", "胃穿孔", "胆结石", "肾结石", "输尿管结石", "前列腺增生", "静脉曲张", "颈椎病", "腰椎间盘", "肩周炎", "股骨头", "半月板", "腱鞘炎"],
    "妇产科": ["月经", "痛经", "闭经", "白带", "阴道", "宫颈", "子宫", "卵巢", "输卵管", "妊娠", "流产", "分娩", "产后", "乳腺", "乳房", "更年期", "盆腔炎", "多囊卵巢"],
    "儿科": ["小儿", "儿童", "婴儿", "幼儿", "新生儿", "早产", "发育", "生长", "接种", "出牙", "腹泻脱水", "惊厥", "高热惊厥", "脑瘫", "先天", "遗尿"],
    "神经内科": ["头痛", "偏头痛", "眩晕", "癫痫", "抽搐", "帕金森", "震颤", "痴呆", "记忆力", "失眠", "嗜睡", "面瘫", "三叉神经", "坐骨神经", "肌无力", "肌萎缩", "共济失调", "脑梗", "脑出血", "蛛网膜", "脑膜炎", "颅内", "脑血管"],
    "骨科": ["骨折", "脱位", "扭伤", "脊柱", "颈椎", "腰椎", "关节", "半月板", "韧带", "肌腱", "滑膜", "骨刺", "骨质增生", "骨质疏松", "骨肿瘤", "骨髓炎", "骨坏死", "股骨头坏死", "椎管狭窄", "滑脱", "侧弯", "畸形", "假体"],
    "心血管内科": ["高血压", "冠心病", "心梗", "心绞痛", "心律失常", "房颤", "室早", "心动过速", "心动过缓", "心包炎", "心肌炎", "心肌病", "瓣膜", "主动脉", "周围血管", "深静脉血栓", "动脉硬化", "高血脂"],
    "呼吸内科": ["咳嗽", "咳痰", "咯血", "气喘", "呼吸困难", "胸闷", "胸痛", "鼻炎", "咽炎", "喉炎", "支气管炎", "肺炎", "肺气肿", "慢阻肺", "哮喘", "肺纤维化", "肺结核", "肺癌", "胸膜", "打鼾", "睡眠呼吸暂停"],
    "消化内科": ["胃痛", "腹痛", "腹泻", "便秘", "恶心", "呕吐", "反酸", "烧心", "嗳气", "腹胀", "消化不良", "胃溃疡", "胃炎", "胃癌", "肝炎", "脂肪肝", "肝硬化", "肝癌", "胆囊", "胰腺炎", "肠炎", "结肠炎", "克罗恩", "肠易激", "便血", "黑便"],
    "耳鼻喉科": ["耳", "听力", "耳鸣", "中耳炎", "鼻塞", "流涕", "鼻出血", "鼻炎", "鼻窦炎", "咽痛", "咽炎", "喉炎", "声带", "声音嘶哑", "扁桃体", "打鼾", "眩晕", "梅尼埃"],
    "眼科": ["眼", "视力", "模糊", "复视", "眼红", "眼痛", "眼干", "眼涩", "畏光", "流泪", "结膜炎", "角膜炎", "白内障", "青光眼", "视网膜", "黄斑", "玻璃体", "近视", "远视", "散光"],
    "皮肤科": ["皮疹", "瘙痒", "红斑", "丘疹", "水疱", "脓疱", "风团", "鳞屑", "脱发", "多汗", "狐臭", "痤疮", "湿疹", "皮炎", "荨麻疹", "银屑病", "带状疱疹", "癣", "痣", "斑", "疣", "痱子"],
    "口腔科": ["牙", "口腔", "牙龈", "智齿", "龋齿", "牙痛", "牙周", "颌", "种植", "正畸", "根管", "溃疡", "出血", "肿胀", "牙髓", "口臭", "唇", "舌", "颊", "黏膜"],
}


def _guess_department(case_text: str) -> str:
    """Heuristic: score case text against keyword dictionaries, return best match."""
    best_dept = "内科"  # default
    best_score = 0
    for dept, keywords in DEPARTMENT_KEYWORDS.items():
        score = 0
        for kw in keywords:
            if kw in case_text:
                score += 1
        if score > best_score:
            best_score = score
            best_dept = dept
    return best_dept


def _difficulty(text: str) -> int:
    """1-3 based on text length."""
    length = len(text)
    if length < 400:
        return 1
    elif length < 800:
        return 2
    return 3


class CaseService:
    """Handles case listing, preview, and patient agent initialization."""

    XLSX_PATH = Path("dataset/patient_text.xlsx")

    _cache: list[dict] | None = None  # lazy-loaded cache

    @classmethod
    def _load_all_cases(cls) -> list[dict]:
        """Read the Excel and return all case summaries (cached)."""
        if cls._cache is not None:
            return cls._cache

        xlsx_path = cls.XLSX_PATH
        if not xlsx_path.is_file():
            cls._cache = []
            return cls._cache

        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        sheet_name = wb.sheetnames[0]
        sheet = wb[sheet_name]

        cases = []
        for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=0):
            if row_idx == 0 or row_idx == 1:  # skip header rows
                continue
            if row[0] is None:
                continue
            # Look for text content in columns (skip the first column which is SN)
            text_parts = []
            for cell in row[1:]:
                if cell is not None:
                    text = str(cell).strip()
                    if text and len(text) > 5:
                        text_parts.append(text)
            full_text = " ".join(text_parts)
            if not full_text:
                continue

            preview = full_text[:200]
            department = _guess_department(full_text)
            diff = _difficulty(full_text)

            cases.append({
                "id": row_idx,
                "serial_number": str(row[0]) if row[0] else f"case-{row_idx}",
                "department": department,
                "preview": preview,
                "difficulty": diff,
                "case_length": len(full_text),
                "practice_count": 0,  # will be populated from DB in Phase 3
            })

        wb.close()
        cls._cache = cases
        return cls._cache

    @classmethod
    def list_cases(
        cls,
        department: str = "",
        difficulty: str = "",
        search: str = "",
        page: int = 1,
        page_size: int = 20,
    ) -> tuple[list[dict], int]:
        """Paginated and filtered case list."""
        all_cases = cls._load_all_cases()

        # Filter
        filtered = []
        for c in all_cases:
            if department and department != "全部" and c["department"] != department:
                continue
            if difficulty and difficulty != "全部" and c["difficulty"] != int(difficulty):
                continue
            if search:
                search_lower = search.lower()
                text = (c["preview"] + " " + c["department"] + " " + c["serial_number"]).lower()
                if search_lower not in text:
                    continue
            filtered.append(c)

        total = len(filtered)
        start = (page - 1) * page_size
        end = start + page_size

        return filtered[start:end], total

    @classmethod
    def load_patient(cls, sheet_name: str, row_number: int, col_number: int = 1) -> dict:
        """Initialize a patient agent — wraps simulateflow.init_session().

        Returns internal data including the SessionContext object.
        The caller (router) is responsible for creating the DB Session record
        and registering the context with SessionService.
        """
        from simulateflow import init_session

        ctx = init_session(
            sheet_name=sheet_name,
            row_number=row_number,
            col_number=col_number,
            output_dir="interactive",
        )

        return {
            "case_id": row_number,
            "sheet_name": sheet_name,
            "department": ctx.office,
            "chief_complaint": ctx.main_complaint,
            "_context": ctx,
        }

    @classmethod
    def get_patient_agent(cls, session_id: str):
        """Retrieve the Patient agent by session ID (delegates to SessionService)."""
        from server.services.session_service import SessionService
        return SessionService.get_patient_agent(session_id)

    @classmethod
    def get_session_context(cls, session_id: str):
        """Retrieve the full SessionContext by session ID (delegates to SessionService)."""
        from server.services.session_service import SessionService
        return SessionService.get_context(session_id)
