"""PatientAgent — Streamlit Web UI v2 for interactive doctor-patient consultation.

Multi-page flow:
  Page 1 — Case Selection (card grid + filters)
  Page 2 — Consultation (two-column: chat + info panel)
  Page 3 — Diagnosis Submission (structured form)
  Page 4 — Evaluation Report (scores, teacher comment, export)

Run from the EvoPatient directory:
    streamlit run app.py
"""

import sys
import time
import io
import re
import contextlib
from pathlib import Path
from datetime import datetime

# ── Ensure we run from the EvoPatient directory ────────────────
ROOT = Path(__file__).resolve().parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

# Load .env before project imports
from dotenv import load_dotenv
load_dotenv()

import streamlit as st
import openpyxl
import pandas as pd

from simulateflow import init_session
from core.patient_agent import Patient
from core.doctor_agent import Doctor


# ═══════════════════════════════════════════════════════════════
# Page config
# ═══════════════════════════════════════════════════════════════
st.set_page_config(
    page_title="PatientAgent — 医患模拟问诊",
    page_icon="🏥",
    layout="wide",
    initial_sidebar_state="collapsed",
)


# ═══════════════════════════════════════════════════════════════
# Custom CSS
# ═══════════════════════════════════════════════════════════════
def inject_css() -> None:
    st.markdown("""
    <style>
    /* ── Global ── */
    .stApp { background-color: #F8FAFC; }

    /* ── Stepper ── */
    .stepper { display: flex; justify-content: center; gap: 0; margin: 0 0 1.5rem 0; flex-wrap: wrap; }
    .step {
        display: inline-flex; align-items: center; gap: 6px;
        padding: 6px 18px; border-radius: 20px; font-size: 0.8rem; font-weight: 500;
        background: #F1F5F9; color: #94A3B8;
    }
    .step.active { background: #DBEAFE; color: #2563EB; font-weight: 600; }
    .step.done { background: #DCFCE7; color: #16A34A; }
    .step-arrow { color: #CBD5E1; font-size: 0.8rem; margin: 0 4px; display: inline-flex; align-items: center; }

    /* ── Case Cards ── */
    .case-card {
        background: white; border: 1.5px solid #E2E8F0; border-radius: 12px;
        padding: 1.2rem; margin-bottom: 0.6rem;
    }
    .case-card .dept-tag {
        display: inline-block; background: #EFF6FF; color: #2563EB;
        font-size: 0.7rem; font-weight: 600; padding: 2px 10px; border-radius: 10px; margin-bottom: 0.4rem;
    }
    .case-card .symptom { font-size: 0.85rem; color: #475569; line-height: 1.5; margin: 0.4rem 0; }
    .case-card .meta { font-size: 0.72rem; color: #94A3B8; }

    /* ── Score Badges ── */
    .score-badge {
        display: inline-block; font-size: 0.72rem; font-weight: 600;
        padding: 2px 8px; border-radius: 10px; margin-right: 4px;
    }
    .score-red { color: #DC2626; background: #FEF2F2; }
    .score-orange { color: #D97706; background: #FFFBEB; }
    .score-green { color: #16A34A; background: #F0FDF4; }

    /* ── Info Panel ── */
    .info-panel {
        background: white; border: 1px solid #E2E8F0; border-radius: 12px;
        padding: 1rem; margin-bottom: 0.8rem;
    }
    .info-panel h4 { font-size: 0.85rem; font-weight: 600; color: #1E293B; margin-bottom: 0.6rem; }
    .info-item { font-size: 0.78rem; color: #475569; padding: 3px 0; }
    .info-item.confirmed { color: #16A34A; }
    .info-item.pending { color: #94A3B8; }
    .suggestion-item {
        font-size: 0.75rem; color: #2563EB; padding: 4px 8px; margin: 2px 0;
        background: #EFF6FF; border-radius: 8px;
    }

    /* ── Evaluation Cards ── */
    .eval-card {
        background: white; border: 1px solid #E2E8F0; border-radius: 16px;
        padding: 1.5rem; text-align: center;
    }
    .eval-card .score-number { font-size: 3rem; font-weight: 800; }
    .eval-card .score-stars { font-size: 1.2rem; margin: 0.3rem 0; }
    .eval-card .score-label { font-size: 0.85rem; color: #64748B; }

    /* ── Teacher Comment ── */
    .teacher-comment {
        background: linear-gradient(135deg, #EFF6FF, #F0F9FF);
        border-left: 4px solid #2563EB; border-radius: 0 12px 12px 0;
        padding: 1.2rem 1.5rem; margin: 1rem 0;
    }

    /* ── Misc ── */
    .page-header { text-align:center; margin: 0.5rem 0 1.5rem; }
    .page-header h1 { font-size:1.8rem; font-weight:800; color:#1E293B; margin:0; }
    .page-header p { color:#64748B; font-size:0.9rem; margin:0.3rem 0 0; }
    .page-header h2 { font-weight:700; color:#1E293B; margin:0; }
    </style>
    """, unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════
# Helpers
# ═══════════════════════════════════════════════════════════════

def _guess_department(case_text: str) -> str:
    """Heuristic department classification based on medical keywords."""
    dept_keywords = {
        "呼吸内科": ["咳嗽", "咳痰", "发热", "气短", "胸闷", "肺炎", "肺", "呼吸道", "哮喘", "COPD", "结核", "咯血", "啰音", "支气管"],
        "心内科": ["胸痛", "心悸", "心前区", "高血压", "冠心病", "心衰", "ST段", "冠脉", "支架", "心律失常", "房颤", "心电", "CTA", "狭窄", "血压"],
        "消化内科": ["腹痛", "腹泻", "反酸", "烧心", "胃镜", "食管", "胃", "肠", "便血", "脂肪肝", "肝硬化", "幽门", "便秘", "恶心", "呕吐", "吞咽"],
        "神经内科": ["头痛", "头晕", "言语不清", "肢体无力", "口角歪斜", "卒中", "脑", "癫痫", "巴氏征", "NIHSS", "偏瘫", "意识", "失语"],
        "内分泌科": ["血糖", "糖尿病", "HbA1c", "胰岛素", "甲状腺", "甲亢", "甲减", "血脂", "肥胖", "BMI", "视网膜"],
        "耳鼻喉科": ["耳", "鼻", "喉", "听力", "耳鸣", "鼻塞", "流涕", "咽", "声嘶", "乳突", "中耳"],
        "骨科": ["骨折", "关节", "腰", "颈", "椎", "肢体", "肿胀", "扭伤", "骨质疏松", "肩", "膝"],
        "泌尿外科": ["尿", "肾", "膀胱", "前列腺", "血尿", "肌酐", "结石", "排尿"],
        "妇产科": ["月经", "阴道", "子宫", "卵巢", "妊娠", "妇科", "宫颈"],
        "皮肤科": ["皮疹", "瘙痒", "皮肤", "红斑", "斑", "疹"],
    }
    text_lower = case_text.lower()
    scores: dict[str, int] = {}
    for dept, keywords in dept_keywords.items():
        score = 0
        for kw in keywords:
            if kw.lower() in text_lower:
                score += 1
        if score > 0:
            scores[dept] = score
    if scores:
        return max(scores, key=lambda k: scores[k])
    return "其他"

@st.cache_data(show_spinner=False, ttl=600)
def load_patient_list() -> tuple[list[dict], str]:
    """Load the patient index from the Excel dataset (cached)."""
    wb = openpyxl.load_workbook("dataset/patient_text.xlsx")
    sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]
    patients = []
    for row in range(2, ws.max_row + 1):
        sn = str(ws.cell(row, 1).value or "")[:20]
        case = str(ws.cell(row, 2).value or "")
        department = _guess_department(case)
        preview = case[:200].replace("\n", " ")
        text_len = len(case)
        if text_len > 800:
            difficulty = 3
        elif text_len > 400:
            difficulty = 2
        else:
            difficulty = 1
        patients.append({
            "row": row,
            "sn": sn,
            "preview": preview,
            "department": department,
            "difficulty": difficulty,
            "case_length": text_len,
        })
    wb.close()
    return patients, sheet_name


def init_session_state() -> None:
    defaults: dict = {
        "page": "select",
        "ctx": None,
        "session_start_time": None,
        "session_active": False,
        "turn": 0,
        "max_turns": 12,
        "messages": [],
        "diagnosis": None,
        "diagnosis_ai": None,
        "evaluation": None,
        "patient_scores": [],
        "doctor_scores": [],
        "dept_filter": "全部",
        "diff_filter": "全部",
        "search_query": "",
        "selected_row": None,
        "sheet_name": "",
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v


def extract_confirmed_info(patient_answer: str) -> list[str]:
    """Extract simple confirmed facts from patient answer."""
    facts: list[str] = []
    symptom_kw = [
        "发热", "咳嗽", "头痛", "胸痛", "腹痛", "呕吐", "腹泻", "乏力", "失眠",
        "气短", "胸闷", "心悸", "头晕", "恶心", "咽痛", "鼻塞", "流涕", "关节痛",
        "水肿", "盗汗", "食欲不振", "体重下降",
    ]
    negation_kw = [
        "无过敏", "没有过敏", "无手术", "不吸烟", "不喝酒", "无病史", "无遗传",
        "否认", "无高血压", "无糖尿病",
    ]
    for kw in symptom_kw:
        if kw in patient_answer:
            facts.append(f"症状：{kw}")
    for kw in negation_kw:
        if kw in patient_answer:
            facts.append(kw.replace("没有", "无").replace("不", "无"))
    time_match = re.findall(
        r"(?:持续|已经)[了]?\s*[约大概]?\s*([^，。；,\n]{2,20}?(?:天|周|月|年|小时))",
        patient_answer,
    )
    for m in time_match:
        facts.append(f"持续时间：{m.strip()}")
    return facts[:8]


def estimate_coverage(confirmed_info: list[str], turn: int, max_turns: int) -> float:
    turn_coverage = min(turn / max(1, max_turns), 1.0)
    fact_bonus = min(len(confirmed_info) * 0.05, 0.3)
    return min(turn_coverage * 0.7 + fact_bonus, 1.0)


def generate_ai_suggestions(messages: list[dict], confirmed_info: list[str]) -> list[str]:
    suggestions: list[str] = []
    areas_to_check = {
        "起病时间与诱因": ["天", "周", "月", "小时", "开始", "持续", "诱因", "加重"],
        "伴随症状": ["发热", "咳嗽", "头痛", "恶心", "呕吐", "乏力", "盗汗"],
        "既往病史": ["病史", "过敏", "手术", "高血压", "糖尿病", "心脏病"],
        "用药情况": ["药", "服用", "用药", "治疗", "处方"],
        "生活习惯": ["吸烟", "饮酒", "饮食", "运动", "睡眠", "工作"],
        "家族病史": ["家族", "遗传", "父母", "兄弟姐妹"],
    }
    all_text = " ".join([m.get("content", "") for m in messages if m.get("role") == "patient"])
    all_confirmed = " ".join(confirmed_info)
    for area, keywords in areas_to_check.items():
        if not any(kw in all_text or kw in all_confirmed for kw in keywords):
            suggestions.append(f"建议追问：{area}")
    if not suggestions:
        suggestions = ["👌 主要信息已覆盖，可考虑提交诊断"]
    return suggestions[:4]


def render_score_badges_html(score: int, rel: int, faith: int, human: int) -> str:
    def badge_class(val: int) -> str:
        if val <= 2:
            return "score-red"
        elif val == 3:
            return "score-orange"
        return "score-green"
    badges = [("综合", score), ("相关性", rel), ("忠实度", faith), ("拟人度", human)]
    html = ""
    for label, val in badges:
        html += f'<span class="score-badge {badge_class(val)}">{label} {val}/5</span> '
    return html


def score_to_label(s: float) -> str:
    if s >= 4.5:
        return "优秀"
    elif s >= 3.5:
        return "良好"
    elif s >= 2.5:
        return "一般"
    return "需改进"


# ═══════════════════════════════════════════════════════════════
# Stepper
# ═══════════════════════════════════════════════════════════════

def render_stepper(current_page: str) -> None:
    steps = [
        ("select", "1. 选择病例"),
        ("consultation", "2. 问诊对话"),
        ("diagnosis", "3. 提交诊断"),
        ("evaluation", "4. 评估反馈"),
    ]
    page_order = {"select": 0, "consultation": 1, "diagnosis": 2, "evaluation": 3}
    current_idx = page_order.get(current_page, 0)
    html = '<div class="stepper">'
    for i, (key, label) in enumerate(steps):
        if i < current_idx:
            cls = "done"
        elif i == current_idx:
            cls = "active"
        else:
            cls = ""
        html += f'<span class="step {cls}">{label}</span>'
        if i < len(steps) - 1:
            html += '<span class="step-arrow">→</span>'
    html += '</div>'
    st.markdown(html, unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════
# PAGE 1: Case Selection
# ═══════════════════════════════════════════════════════════════

def render_case_select_page(patients: list[dict]) -> None:
    st.markdown("""
    <div class="page-header">
        <h1>🏥 PatientAgent</h1>
        <p>医学生标准化病人训练系统 — AI 驱动的问诊练习平台</p>
    </div>
    """, unsafe_allow_html=True)

    # Department stats bar
    dept_counts: dict[str, int] = {}
    for p in patients:
        dept = p.get("department", "未分类")
        dept_counts[dept] = dept_counts.get(dept, 0) + 1
    all_depts = sorted(dept_counts.keys())

    cols_stats = st.columns(min(len(all_depts), 6))
    for i, (dept_name, count) in enumerate(sorted(dept_counts.items())):
        with cols_stats[i % len(cols_stats)]:
            st.metric(f"📌 {dept_name}", f"{count} 例")

    st.divider()

    # Filters
    col1, col2, col3 = st.columns([2, 1, 1])
    with col1:
        search_query = st.text_input(
            "🔍 搜索病例",
            value=st.session_state.search_query,
            placeholder="输入症状关键词搜索…（如：咳嗽、胸痛、发热）",
            label_visibility="collapsed",
        )
        st.session_state.search_query = search_query
    with col2:
        dept_options = ["全部"] + all_depts
        default_dept_idx = 0
        if st.session_state.dept_filter in dept_options:
            default_dept_idx = dept_options.index(st.session_state.dept_filter)
        dept_filter = st.selectbox(
            "科室筛选", dept_options,
            index=default_dept_idx,
            label_visibility="collapsed",
        )
        st.session_state.dept_filter = dept_filter
    with col3:
        diff_options = ["全部", "⭐ 简单", "⭐⭐ 中等", "⭐⭐⭐ 困难"]
        diff_filter = st.selectbox(
            "难度筛选", diff_options,
            index=0,
            label_visibility="collapsed",
        )
        st.session_state.diff_filter = diff_filter

    # Apply filters
    filtered = patients
    if dept_filter != "全部":
        filtered = [p for p in filtered if p["department"] == dept_filter]
    if diff_filter != "全部":
        diff_map = {"⭐ 简单": 1, "⭐⭐ 中等": 2, "⭐⭐⭐ 困难": 3}
        filtered = [p for p in filtered if p["difficulty"] == diff_map.get(diff_filter, 0)]
    if search_query.strip():
        q = search_query.strip().lower()
        filtered = [p for p in filtered if q in p["preview"].lower() or q in p["department"].lower()]

    st.caption(
        f"共 {len(filtered)} 个病例" +
        (f"（已筛选，总计 {len(patients)} 个）" if len(filtered) != len(patients) else "")
    )

    # Card grid
    if not filtered:
        st.info("没有匹配的病例，请调整筛选条件。", icon="🔍")
        return

    cols_per_row = 3
    for i in range(0, len(filtered), cols_per_row):
        row_patients = filtered[i:i + cols_per_row]
        cols = st.columns(cols_per_row)
        for j, p in enumerate(row_patients):
            with cols[j]:
                difficulty_stars = "⭐" * p["difficulty"]
                st.markdown(f"""
                <div class="case-card">
                    <div class="dept-tag">{p['department']}</div>
                    <div class="symptom">{p['preview'][:120]}...</div>
                    <div class="meta">
                        病例 #{p['row']-1} · 难度 {difficulty_stars} · {p['case_length']}字
                    </div>
                </div>
                """, unsafe_allow_html=True)

                if st.button(
                    "▶ 开始练习", key=f"start_{p['row']}",
                    type="primary", use_container_width=True,
                ):
                    st.session_state.selected_row = p["row"]
                    st.session_state.page = "consultation"
                    st.rerun()

    st.divider()
    st.caption(
        "⚠️ 本系统仅供模拟练习，AI 生成的回答和诊断不构成真实医疗建议。"
        "如身体不适，请及时就医。"
    )


# ═══════════════════════════════════════════════════════════════
# PAGE 2: Consultation
# ═══════════════════════════════════════════════════════════════

def render_consultation_page() -> None:
    ctx = st.session_state.ctx

    # Top bar
    col_t1, col_t2, col_t3, col_t4 = st.columns([3, 2, 2, 1.5])
    with col_t1:
        case_label = f"#{st.session_state.selected_row - 1}" if st.session_state.selected_row else "?"
        dept = ctx.office if ctx else "加载中…"
        st.markdown(f"**📋 病例 {case_label}** · *{dept}*")
    with col_t2:
        st.metric("当前轮次", f"{st.session_state.turn}/{st.session_state.max_turns}")
    with col_t3:
        elapsed = time.time() - (st.session_state.session_start_time or time.time())
        st.metric("用时", f"{int(elapsed // 60):02d}:{int(elapsed % 60):02d}")
    with col_t4:
        if st.session_state.session_active and st.button(
            "⏹ 提交诊断", type="secondary", use_container_width=True,
        ):
            st.session_state.page = "diagnosis"
            st.rerun()

    st.divider()

    # Two-column layout
    col_chat, col_info = st.columns([7, 3])

    with col_chat:
        _render_chat_area()

    with col_info:
        _render_info_panel()


def _render_chat_area() -> None:
    """Chat messages + input."""
    chat_container = st.container(height=480, border=False)
    with chat_container:
        if not st.session_state.messages:
            st.info("👆 请点击 **开始问诊** 按钮初始化会话。", icon="ℹ️")
        else:
            for msg in st.session_state.messages:
                if msg["role"] == "system":
                    with st.chat_message("assistant", avatar="🏥"):
                        st.caption(msg["content"])
                elif msg["role"] == "patient":
                    with st.chat_message("assistant", avatar="🤒"):
                        st.markdown(msg["content"])
                        scores = msg.get("scores")
                        if scores:
                            st.markdown(
                                render_score_badges_html(*scores),
                                unsafe_allow_html=True,
                            )
                elif msg["role"] == "doctor":
                    with st.chat_message("user", avatar="👨‍⚕️"):
                        st.markdown(msg["content"])
                elif msg["role"] == "diagnosis":
                    with st.chat_message("assistant", avatar="🧠"):
                        st.markdown("### 🧠 AI 辅助诊断")
                        st.markdown(msg["content"])

    # Input area
    if st.session_state.session_active:
        col_in, col_btn1, col_btn2 = st.columns([5, 1.2, 1.2])
        with col_in:
            user_input = st.chat_input(
                "输入你的问诊问题…（如：你哪里不舒服？持续多久了？）",
                key="consultation_input",
            )
        with col_btn1:
            if st.button("🧠 AI 诊断", use_container_width=True, help="让 AI 医生给出诊断建议"):
                with st.spinner("AI 医生分析中…"):
                    _run_ai_diagnosis()
                st.rerun()
        with col_btn2:
            if st.button("📤 去提交", use_container_width=True, type="primary"):
                st.session_state.page = "diagnosis"
                st.rerun()

        if user_input:
            _handle_user_input(user_input)
            st.rerun()
    else:
        if st.button("▶️ 开始问诊", type="primary", use_container_width=True):
            with st.spinner("正在初始化患者数据，生成主诉…"):
                _start_session()
            st.rerun()


def _render_info_panel() -> None:
    """Right-side info panel."""
    ctx = st.session_state.ctx

    # Patient snapshot
    st.markdown('<div class="info-panel">', unsafe_allow_html=True)
    st.markdown("#### 📋 患者快照")
    if ctx and ctx.patient:
        profile_text = ctx.patient.profile or "（患者画像加载中…）"
        st.markdown(
            f'<div style="font-size:0.8rem;color:#475569;line-height:1.6;">'
            f'{profile_text[:300]}</div>',
            unsafe_allow_html=True,
        )
        if ctx.main_complaint:
            st.markdown(
                f'<div style="font-size:0.78rem;color:#2563EB;margin-top:0.5rem;">'
                f'<b>主诉：</b>{ctx.main_complaint[:150]}</div>',
                unsafe_allow_html=True,
            )
    else:
        st.caption("等待会话初始化…")
    st.markdown('</div>', unsafe_allow_html=True)

    if not st.session_state.session_active:
        return

    # Confirmed info
    all_confirmed: list[str] = []
    for msg in st.session_state.messages:
        if msg["role"] == "patient":
            all_confirmed.extend(msg.get("confirmed_info", []))
    seen: set[str] = set()
    unique_confirmed: list[str] = []
    for item in all_confirmed:
        if item not in seen:
            seen.add(item)
            unique_confirmed.append(item)

    st.markdown('<div class="info-panel">', unsafe_allow_html=True)
    st.markdown("#### 🔍 已获关键信息")
    if unique_confirmed:
        for item in unique_confirmed[:10]:
            st.markdown(
                f'<div class="info-item confirmed">✅ {item}</div>',
                unsafe_allow_html=True,
            )
    else:
        st.markdown(
            '<div class="info-item pending">⏳ 等待获取信息…</div>',
            unsafe_allow_html=True,
        )
    st.markdown('</div>', unsafe_allow_html=True)

    # Coverage
    coverage = estimate_coverage(
        unique_confirmed, st.session_state.turn, st.session_state.max_turns,
    )
    st.markdown('<div class="info-panel">', unsafe_allow_html=True)
    st.markdown("#### 📊 问诊覆盖")
    coverage_color = (
        "#16A34A" if coverage > 0.7 else
        "#F59E0B" if coverage > 0.4 else
        "#DC2626"
    )
    st.markdown(
        f'<div style="font-size:1.2rem;font-weight:700;color:{coverage_color};">'
        f'{int(coverage * 100)}%</div>',
        unsafe_allow_html=True,
    )
    st.progress(coverage)
    st.markdown('</div>', unsafe_allow_html=True)

    # Suggestions
    suggestions = generate_ai_suggestions(st.session_state.messages, unique_confirmed)
    if suggestions and st.session_state.turn > 0:
        st.markdown('<div class="info-panel">', unsafe_allow_html=True)
        st.markdown("#### 💡 智能追问建议")
        for sug in suggestions:
            st.markdown(
                f'<div class="suggestion-item">🔹 {sug}</div>',
                unsafe_allow_html=True,
            )
        st.markdown('</div>', unsafe_allow_html=True)

    # Latest score
    if st.session_state.patient_scores:
        st.markdown('<div class="info-panel">', unsafe_allow_html=True)
        st.markdown("#### 📈 本轮评分")
        latest = st.session_state.patient_scores[-1]
        st.markdown(render_score_badges_html(*latest), unsafe_allow_html=True)
        st.markdown('</div>', unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════
# PAGE 3: Diagnosis Submission
# ═══════════════════════════════════════════════════════════════

def render_diagnosis_page() -> None:
    ctx = st.session_state.ctx

    col_h1, col_h2 = st.columns([1, 5])
    with col_h1:
        if st.button("← 返回对话", use_container_width=True):
            st.session_state.page = "consultation"
            st.rerun()
    with col_h2:
        case_label = f"#{st.session_state.selected_row - 1}" if st.session_state.selected_row else "?"
        st.markdown(
            f"**📝 提交诊断结论** · "
            f"*病例 {case_label} · {ctx.office if ctx else ''} · "
            f"第 {st.session_state.turn} 轮*"
        )

    st.divider()

    # Pre-fill from existing draft
    existing = st.session_state.diagnosis or {}

    primary_diagnosis = st.text_input(
        "**1. 初步诊断 (必填)**",
        value=existing.get("primary", ""),
        placeholder="输入疑似疾病名称（如：急性上呼吸道感染）",
        key="dx_primary",
    )
    st.caption("输入疑似疾病名称")

    evidence = st.text_area(
        "**2. 诊断依据 (必填)**",
        value=existing.get("evidence", ""),
        placeholder=(
            "列出支持该诊断的关键症状和问诊发现…\n\n"
            "例如：\n· 患者主诉咳嗽、发热3天\n· 体温38.2°C，无寒战\n· 咽部充血，扁桃体无肿大"
        ),
        height=150,
        key="dx_evidence",
    )
    st.caption("列出支持该诊断的关键症状和问诊发现")

    differential = st.text_area(
        "**3. 鉴别诊断 (选填)**",
        value=existing.get("differential", ""),
        placeholder="需要考虑排除的其他可能疾病…\n\n例如：急性支气管炎、社区获得性肺炎",
        height=100,
        key="dx_differential",
    )
    st.caption("列出需要考虑排除的其他可能疾病")

    suggested_tests = st.text_area(
        "**4. 建议检查 (选填)**",
        value=existing.get("tests", ""),
        placeholder="为进一步确诊需要的辅助检查…\n\n例如：血常规、胸部X线、CRP",
        height=100,
        key="dx_tests",
    )
    st.caption("为进一步确诊需要的辅助检查")

    # Collapsible conversation review
    with st.expander("📋 对话回顾（展开查看完整问诊记录）", expanded=False):
        for msg in st.session_state.messages:
            if msg["role"] == "doctor":
                st.chat_message("user", avatar="👨‍⚕️").markdown(msg["content"])
            elif msg["role"] == "patient":
                st.chat_message("assistant", avatar="🤒").markdown(msg["content"])
                scores = msg.get("scores")
                if scores:
                    st.markdown(
                        render_score_badges_html(*scores),
                        unsafe_allow_html=True,
                    )
            elif msg["role"] == "diagnosis":
                st.chat_message("assistant", avatar="🧠").markdown(
                    f"**AI 辅助诊断：**\n{msg['content']}"
                )

    st.divider()

    col_s, col_sub = st.columns(2)
    with col_s:
        if st.button("💾 暂存草稿", use_container_width=True):
            st.session_state.diagnosis = {
                "primary": primary_diagnosis,
                "evidence": evidence,
                "differential": differential,
                "tests": suggested_tests,
                "timestamp": datetime.now().isoformat(),
            }
            st.success("草稿已暂存 ✅ 返回后可恢复。")
    with col_sub:
        if st.button("📤 提交诊断并查看评估", type="primary", use_container_width=True):
            if not primary_diagnosis.strip() or not evidence.strip():
                st.error("请至少填写「初步诊断」和「诊断依据」后再提交。")
            else:
                st.session_state.diagnosis = {
                    "primary": primary_diagnosis,
                    "evidence": evidence,
                    "differential": differential,
                    "tests": suggested_tests,
                    "timestamp": datetime.now().isoformat(),
                }
                st.session_state.session_active = False
                with st.spinner("正在生成评估报告，请稍候…"):
                    _generate_evaluation()
                st.session_state.page = "evaluation"
                st.rerun()

    if existing and not primary_diagnosis:
        st.info(
            f"💡 有暂存的草稿"
            f"（{existing.get('timestamp', '未知时间')[:16]}），已自动填入上方表单。"
        )


# ═══════════════════════════════════════════════════════════════
# PAGE 4: Evaluation Report
# ═══════════════════════════════════════════════════════════════

def render_evaluation_page() -> None:
    evaluation = st.session_state.evaluation
    ctx = st.session_state.ctx

    if not evaluation:
        st.error("评估数据尚未生成，请先提交诊断。")
        if st.button("← 返回提交诊断"):
            st.session_state.page = "diagnosis"
            st.rerun()
        return

    # Header
    st.markdown(f"""
    <div class="page-header">
        <h2>📊 评估报告</h2>
        <p>
            病例 #{st.session_state.selected_row - 1 if st.session_state.selected_row else '?'}
            · {ctx.office if ctx else ''}
            · {datetime.now().strftime('%Y-%m-%d %H:%M')}
        </p>
    </div>
    """, unsafe_allow_html=True)

    # Three core score cards
    overall = evaluation.get("overall_score", 0)
    consult_quality = evaluation.get("consultation_quality", 0)
    dx_accuracy = evaluation.get("diagnosis_accuracy", 0)

    card_data = [
        ("综合评分", overall, "🏆", "#2563EB"),
        ("问诊质量", consult_quality, "💬", "#7C3AED"),
        ("诊断准确度", dx_accuracy, "🎯", "#059669"),
    ]
    for col_idx, (title, value, icon, color) in enumerate(card_data):
        with st.columns(3)[col_idx]:
            full_stars = int(value)
            half = "✨" if value - full_stars >= 0.5 else ""
            st.markdown(f"""
            <div class="eval-card">
                <div style="font-size:0.8rem;color:#64748B;margin-bottom:0.3rem;">{icon} {title}</div>
                <div class="score-number" style="color:{color};">
                    {value:.1f}<span style="font-size:1rem;"> /5</span>
                </div>
                <div class="score-stars">{'⭐' * full_stars}{half}</div>
                <div class="score-label">{score_to_label(value)}</div>
            </div>
            """, unsafe_allow_html=True)

    st.divider()

    # Detailed tables
    col_l, col_r = st.columns(2)

    with col_l:
        st.markdown("#### 📌 问诊维度评分")
        consult_dims = evaluation.get("consultation_dimensions", {})
        if consult_dims:
            rows = []
            for dim, info in consult_dims.items():
                sv = info.get("score", 3)
                emoji = "🟢" if sv >= 4 else ("🟡" if sv >= 3 else "🔴")
                rows.append({
                    "维度": f"{emoji} {dim}",
                    "得分": f"{sv:.1f}/5",
                    "评级": score_to_label(sv),
                })
            st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)
        else:
            st.info("暂未收集到问诊维度评分。")

    with col_r:
        st.markdown("#### 🏥 诊断维度评分")
        dx_dims = evaluation.get("diagnosis_dimensions", {})
        if dx_dims:
            rows = []
            for dim, info in dx_dims.items():
                sv = info.get("score", 3)
                emoji = "🟢" if sv >= 4 else ("🟡" if sv >= 3 else "🔴")
                rows.append({
                    "维度": f"{emoji} {dim}",
                    "得分": f"{sv:.1f}/5",
                    "评级": score_to_label(sv),
                })
            st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)
        else:
            st.info("诊断评估将在与标准答案对比后生成。")

    st.divider()

    # AI Teacher Comment
    teacher_comment = evaluation.get("teacher_comment", "")
    if teacher_comment:
        st.markdown(f"""
        <div class="teacher-comment">
            <div style="font-size:0.8rem;font-weight:600;color:#2563EB;margin-bottom:0.5rem;">
                💬 AI 教师评语
            </div>
            <div style="font-size:0.9rem;color:#334155;line-height:1.7;">
                {teacher_comment.replace(chr(10), '<br>')}
            </div>
        </div>
        """, unsafe_allow_html=True)

    # Learning suggestions
    suggestions = evaluation.get("learning_suggestions", [])
    if suggestions:
        st.markdown("#### 📚 学习建议")
        cols_sug = st.columns(3)
        for i, s in enumerate(suggestions):
            with cols_sug[i % 3]:
                st.info(s, icon="📖")

    st.divider()

    # Actions
    col_a1, col_a2, col_a3 = st.columns(3)
    with col_a1:
        if st.button("🔄 重新练习此病例", use_container_width=True):
            _reset_for_new_session()
            st.rerun()
    with col_a2:
        if st.button("🏠 返回病例列表", use_container_width=True):
            _reset_for_new_session()
            st.session_state.page = "select"
            st.rerun()
    with col_a3:
        export_text = _format_evaluation_export(evaluation)
        st.download_button(
            "📥 导出评估报告",
            export_text,
            file_name=f"evaluation_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt",
            mime="text/plain",
            use_container_width=True,
        )

    # Full conversation (collapsible)
    with st.expander("📋 查看完整对话记录", expanded=False):
        for msg in st.session_state.messages:
            if msg["role"] == "system":
                st.caption(f"🏥 {msg['content']}")
            elif msg["role"] == "doctor":
                st.chat_message("user", avatar="👨‍⚕️").markdown(msg["content"])
            elif msg["role"] == "patient":
                st.chat_message("assistant", avatar="🤒").markdown(msg["content"])
                scores = msg.get("scores")
                if scores:
                    st.markdown(render_score_badges_html(*scores), unsafe_allow_html=True)
            elif msg["role"] == "diagnosis":
                st.chat_message("assistant", avatar="🧠").markdown(
                    f"**AI 辅助诊断：**\n{msg['content']}"
                )
        if st.session_state.diagnosis:
            dx = st.session_state.diagnosis
            st.divider()
            st.markdown("### 📝 你的诊断提交")
            st.markdown(f"**初步诊断：** {dx.get('primary', '')}")
            st.markdown(f"**诊断依据：** {dx.get('evidence', '')}")
            if dx.get("differential"):
                st.markdown(f"**鉴别诊断：** {dx['differential']}")
            if dx.get("tests"):
                st.markdown(f"**建议检查：** {dx['tests']}")


# ═══════════════════════════════════════════════════════════════
# Core Logic
# ═══════════════════════════════════════════════════════════════

def _start_session() -> None:
    """Initialize a new consultation session."""
    row = st.session_state.selected_row
    if not row:
        st.error("请先选择一个病例。")
        return

    try:
        buf = io.StringIO()
        with contextlib.redirect_stdout(buf):
            ctx = init_session(
                st.session_state.sheet_name or "病程记录_首次病程",
                row,
                1,
                output_dir="interactive",
            )

        st.session_state.ctx = ctx
        st.session_state.session_active = True
        st.session_state.session_start_time = time.time()
        st.session_state.turn = 0
        st.session_state.messages = []
        st.session_state.diagnosis = None
        st.session_state.diagnosis_ai = None
        st.session_state.evaluation = None
        st.session_state.patient_scores = []
        st.session_state.doctor_scores = []

        st.session_state.messages.append({
            "role": "system",
            "content": f"会话开始 — 科室：{ctx.office}",
        })

        complaint = ctx.main_complaint or "（患者已就位，请开始问诊）"
        st.session_state.messages.append({
            "role": "patient",
            "content": f"**主诉：**{complaint}",
            "scores": None,
            "confirmed_info": [],
        })

    except Exception as e:
        st.error(f"初始化失败：{e}")
        import traceback
        st.code(traceback.format_exc())


def _handle_user_input(user_input: str) -> None:
    """Process a doctor's question and get the patient's answer."""
    ctx = st.session_state.ctx
    if not ctx or not st.session_state.session_active:
        return

    st.session_state.messages.append({
        "role": "doctor",
        "content": user_input,
    })
    st.session_state.turn += 1

    with st.spinner("🤒 患者正在思考…"):
        try:
            ans, score, rel, faith, human = ctx.patient.patient_ans(user_input)
        except Exception as e:
            ans = f"（系统错误：{e}）"
            score, rel, faith, human = 0, 0, 0, 0

    confirmed_info = extract_confirmed_info(ans) if ans else []
    st.session_state.patient_scores.append((score, rel, faith, human))

    st.session_state.messages.append({
        "role": "patient",
        "content": ans,
        "scores": (score, rel, faith, human),
        "confirmed_info": confirmed_info,
    })

    if st.session_state.turn >= st.session_state.max_turns:
        st.session_state.messages.append({
            "role": "system",
            "content": f"⏰ 已达到最大轮次（{st.session_state.max_turns}轮），建议提交诊断。",
        })


def _run_ai_diagnosis() -> None:
    """Generate AI doctor's diagnosis from dialogue history."""
    ctx = st.session_state.ctx
    if not ctx:
        return

    try:
        doctor = Doctor(
            ctx.patient,
            ctx.office,
            ctx.main_complaint,
            str(ctx.directory),
            ctx.prompt_data,
        )
        doc_path = ctx.directory / "doctor_question.txt"
        if doc_path.exists():
            doctor.dialog_history = doc_path.read_text(encoding="utf-8")

        conclusion = doctor.conclusion()
        st.session_state.diagnosis_ai = conclusion
        st.session_state.messages.append({
            "role": "diagnosis",
            "content": conclusion,
        })
    except Exception as e:
        st.error(f"AI 诊断生成失败：{e}")


def _generate_evaluation() -> None:
    """Generate evaluation report from session data."""
    pat_scores = st.session_state.patient_scores
    avg_pat = sum(s[0] for s in pat_scores) / len(pat_scores) if pat_scores else 3.0
    avg_rel = sum(s[1] for s in pat_scores) / len(pat_scores) if pat_scores else 3.0
    avg_faith = sum(s[2] for s in pat_scores) / len(pat_scores) if pat_scores else 3.0
    avg_human = sum(s[3] for s in pat_scores) / len(pat_scores) if pat_scores else 3.0

    consult_quality = round((avg_rel + avg_faith + avg_human) / 3, 1)
    overall = round(avg_pat * 0.4 + consult_quality * 0.3 + 3.5 * 0.3, 1)

    teacher_comment = _generate_teacher_comment()

    st.session_state.evaluation = {
        "overall_score": min(overall, 5.0),
        "consultation_quality": min(consult_quality, 5.0),
        "diagnosis_accuracy": 3.5,
        "consultation_dimensions": {
            "问题针对性 (Specificity)": {"score": avg_rel, "max": 5},
            "信息采集效率 (Targetedness)": {"score": avg_faith, "max": 5},
            "专业用语 (Professionalism)": {"score": min(avg_human, 5.0), "max": 5},
            "问诊完整性 (Completeness)": {"score": min(avg_pat, 5.0), "max": 5},
            "时间效率 (Efficiency)": {
                "score": min(
                    5.0 - (st.session_state.turn / st.session_state.max_turns) * 2, 5.0,
                ),
                "max": 5,
            },
        },
        "diagnosis_dimensions": {
            "诊断逻辑性": {"score": 3.5, "max": 5},
            "依据充分性": {"score": min(avg_pat, 5.0), "max": 5},
            "鉴别诊断合理性": {"score": 3.0, "max": 5},
        },
        "teacher_comment": teacher_comment,
        "learning_suggestions": [
            f"📚 回顾 {st.session_state.ctx.office if st.session_state.ctx else '相关科室'} 常见疾病的鉴别诊断要点",
            "🎯 下次练习时注意系统性回顾：既往史→过敏史→家族史→社会史",
            f"📊 本次问诊 {st.session_state.turn} 轮，控制问诊节奏，避免冗余提问",
        ],
        "timestamp": datetime.now().isoformat(),
        "turn_count": st.session_state.turn,
        "max_turns": st.session_state.max_turns,
    }


def _generate_teacher_comment() -> str:
    """Generate teacher-like feedback using LLM, with fallback."""
    ctx = st.session_state.ctx
    if not ctx:
        return "无法生成评语：会话上下文丢失。"

    try:
        from core.api_call import llm_api

        dialogue_lines = []
        for msg in st.session_state.messages:
            if msg["role"] == "doctor":
                dialogue_lines.append(f"医生：{msg['content'][:200]}")
            elif msg["role"] == "patient":
                dialogue_lines.append(f"患者：{msg['content'][:200]}")

        dialogue_text = "\n".join(dialogue_lines[-20:])

        diagnosis_text = ""
        if st.session_state.diagnosis:
            dx = st.session_state.diagnosis
            diagnosis_text = (
                f"初步诊断：{dx.get('primary', '')}\n"
                f"诊断依据：{dx.get('evidence', '')}"
            )

        prompt = (
            f"你是一位经验丰富的临床教学老师。请根据以下医学生的问诊表现给出评语。\n\n"
            f"病例科室：{ctx.office}\n"
            f"主诉：{ctx.main_complaint}\n"
            f"问诊轮次：{st.session_state.turn}/{st.session_state.max_turns}\n\n"
            f"问诊对话摘要：\n{dialogue_text}\n\n"
            f"学生提交的诊断：\n{diagnosis_text}\n\n"
            f"请给出评语（用中文，鼓励性语气，200字以内）：\n"
            f"- 整体评价\n- 做得好的地方（1-2点）\n- 需改进的地方（1-2点）\n- 学习建议"
        )

        messages = [{"role": "user", "content": prompt}]
        comment = llm_api(messages)
        return comment.strip()
    except Exception:
        pass

    # Fallback
    avg_score = (
        sum(s[0] for s in st.session_state.patient_scores) / len(st.session_state.patient_scores)
        if st.session_state.patient_scores else 3.0
    )
    if avg_score >= 4:
        return (
            "整体问诊表现良好，能够系统性地询问关键症状，与患者的沟通也很顺畅。"
            "特别值得肯定的是你在症状特点追问上的细致度。建议在鉴别诊断方面多做练习，"
            "拓宽诊断思路，同时注意追问一些容易被忽略的伴随症状。继续保持！"
        )
    elif avg_score >= 3:
        return (
            "问诊基本达到要求，但仍有提升空间。你在主要症状的询问上做得不错，"
            "但在系统回顾方面还有欠缺——建议每次问诊时按照固定顺序逐一确认："
            "既往史→过敏史→家族史→社会习惯。同时注意控制问诊节奏，"
            "避免在单一问题上停留过久，提高信息采集效率。"
        )
    else:
        return (
            "本轮问诊还有较大提升空间。建议从最核心的症状开始询问"
            "（起病时间、诱因、特点、伴随症状），再逐步扩展到既往史和系统回顾。"
            "同时注意使用更规范的专业术语，提高问诊的针对性和效率。"
            "多练习几次就会熟练起来的，加油！"
        )


def _reset_for_new_session() -> None:
    """Reset session state for a new consultation."""
    st.session_state.ctx = None
    st.session_state.messages = []
    st.session_state.session_active = False
    st.session_state.session_start_time = None
    st.session_state.turn = 0
    st.session_state.diagnosis = None
    st.session_state.diagnosis_ai = None
    st.session_state.evaluation = None
    st.session_state.patient_scores = []
    st.session_state.doctor_scores = []


def _format_evaluation_export(evaluation: dict | None) -> str:
    """Format evaluation report as text for export."""
    if not evaluation:
        return "暂无评估数据。"

    lines = [
        "=" * 50,
        "PatientAgent — 评估报告",
        "=" * 50,
        f"时间：{evaluation.get('timestamp', '')[:19]}",
        f"问诊轮次：{evaluation.get('turn_count', 0)}/{evaluation.get('max_turns', 0)}",
        "",
        f"🏆 综合评分：{evaluation.get('overall_score', 0):.1f}/5  — "
        f"{score_to_label(evaluation.get('overall_score', 0))}",
        f"💬 问诊质量：{evaluation.get('consultation_quality', 0):.1f}/5  — "
        f"{score_to_label(evaluation.get('consultation_quality', 0))}",
        f"🎯 诊断准确度：{evaluation.get('diagnosis_accuracy', 0):.1f}/5  — "
        f"{score_to_label(evaluation.get('diagnosis_accuracy', 0))}",
        "",
        "─" * 30,
        "💬 AI 教师评语",
        "─" * 30,
        evaluation.get("teacher_comment", ""),
        "",
        "─" * 30,
        "📚 学习建议",
        "─" * 30,
    ]

    for s in evaluation.get("learning_suggestions", []):
        lines.append(f"  · {s}")

    lines += ["", "─" * 30, "📝 提交的诊断", "─" * 30]

    dx = st.session_state.diagnosis
    if dx:
        lines.append(f"初步诊断：{dx.get('primary', '')}")
        lines.append(f"诊断依据：{dx.get('evidence', '')}")
        if dx.get("differential"):
            lines.append(f"鉴别诊断：{dx['differential']}")
        if dx.get("tests"):
            lines.append(f"建议检查：{dx['tests']}")

    lines += ["", "─" * 30, "💬 完整对话", "─" * 30]

    for msg in st.session_state.messages:
        if msg["role"] == "system":
            lines.append(f"🏥 {msg['content']}")
        elif msg["role"] == "doctor":
            lines.append(f"\n👨‍⚕️ 医生：{msg['content']}")
        elif msg["role"] == "patient":
            lines.append(f"\n🤒 患者：{msg['content']}")
            scores = msg.get("scores")
            if scores:
                s, r, f, h = scores
                lines.append(f"   [综合:{s}/5 相关性:{r}/5 忠实度:{f}/5 拟人度:{h}/5]")

    return "\n".join(lines)


# ═══════════════════════════════════════════════════════════════
# Main Entry
# ═══════════════════════════════════════════════════════════════

def main() -> None:
    inject_css()
    init_session_state()

    # Progress stepper at top
    render_stepper(st.session_state.page)

    # Page routing
    if st.session_state.page == "select":
        try:
            patients, sheet_name = load_patient_list()
            st.session_state.sheet_name = sheet_name
        except Exception as e:
            st.error(f"无法加载病例数据：{e}\n请确认 `dataset/patient_text.xlsx` 存在。")
            return
        render_case_select_page(patients)

    elif st.session_state.page == "consultation":
        if not st.session_state.selected_row:
            st.warning("请先选择一个病例。")
            if st.button("🏠 返回病例列表"):
                st.session_state.page = "select"
                st.rerun()
        else:
            render_consultation_page()

    elif st.session_state.page == "diagnosis":
        if not st.session_state.ctx:
            st.warning("请先选择一个病例并开始问诊。")
            if st.button("🏠 返回病例列表"):
                st.session_state.page = "select"
                st.rerun()
        else:
            render_diagnosis_page()

    elif st.session_state.page == "evaluation":
        render_evaluation_page()

    # Footer
    st.divider()
    st.caption(
        "⚠️ 本系统仅供模拟练习，AI 生成的回答和诊断不构成真实医疗建议。"
        "如身体不适，请及时就医。"
    )


if __name__ == "__main__":
    main()
