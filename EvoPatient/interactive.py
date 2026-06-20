"""Interactive doctor-patient consultation.

You play the doctor — type questions in natural language.
The AI patient responds based on a real medical record.

Type 'quit' to end, 'diagnosis' to see the AI doctor's conclusion.
"""

import sys
from pathlib import Path

# Load .env before any imports
from dotenv import load_dotenv
load_dotenv()

import openpyxl
from simulateflow import init_session
from core.patient_agent import Patient


def get_excel_sheet_and_max_row():
    """Find the data sheet and return (sheet_name, max_data_row)."""
    wb = openpyxl.load_workbook("dataset/patient_text.xlsx")
    sheet_name = wb.sheetnames[0]
    max_row = wb[sheet_name].max_row
    wb.close()
    return sheet_name, max_row


def pick_patient():
    """Let user pick a patient row from the Excel file."""
    sheet_name, max_row = get_excel_sheet_and_max_row()
    print(f"\nDataset: {max_row - 1} patients available (rows 2-{max_row})")

    while True:
        choice = input(f"Pick a row number [2-{max_row}], default=2: ").strip()
        if choice == "":
            return sheet_name, 2
        try:
            row = int(choice)
            if 2 <= row <= max_row:
                return sheet_name, row
            print(f"  Please enter a number between 2 and {max_row}")
        except ValueError:
            print("  Please enter a valid number")


def main():
    print("=" * 60)
    print("  PatientAgent — Interactive Consultation Mode")
    print("  You are the doctor. The AI is the patient.")
    print("=" * 60)

    # Pick patient
    sheet_name, row_number = pick_patient()

    # Initialize session
    print("\nLoading patient data...")
    ctx = init_session(sheet_name, row_number, 1, output_dir="interactive")

    print(f"\nDepartment: {ctx.office}")
    print(f"Patient profile loaded.\n")
    print("-" * 60)
    print(f"Patient says: \"{ctx.main_complaint}\"")
    print("-" * 60)
    print("\nStart asking questions! (type 'quit' to end, 'diagnosis' for AI conclusion)\n")

    turn = 0
    while True:
        turn += 1
        doctor_question = input(f"[Turn {turn}] You: ").strip()

        if not doctor_question:
            continue

        if doctor_question.lower() == "quit":
            print("\nConsultation ended.")
            break

        if doctor_question.lower() == "diagnosis":
            print("\n--- AI Doctor Diagnosis ---")
            # Use the AI doctor to generate a conclusion
            from core.doctor_agent import Doctor
            doctor = Doctor(ctx.patient, ctx.office, ctx.main_complaint,
                           str(ctx.directory), ctx.prompt_data)
            # Feed the dialogue history
            doc_path = ctx.directory / "doctor_question.txt"
            if doc_path.exists():
                doctor.dialog_history = doc_path.read_text(encoding="utf-8")
            conclusion = doctor.conclusion()
            print(conclusion)
            print("-" * 60)
            continue

        # Patient answers
        answer, score, rel, faith, human = ctx.patient.patient_ans(doctor_question)
        print(f"\nPatient: {answer}\n")
        print(f"  [quality: {score}/5 | relevance: {rel} | faithfulness: {faith} | human-likeness: {human}]")
        print()


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n\nConsultation ended.")
    except Exception as e:
        print(f"\nError: {e}")
        sys.exit(1)
